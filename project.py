import bdclasses
import datetime
import sqlite3
from sqlite3 import Error
from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Color, PatternFill, Font, Border, colors
from openpyxl.cell import Cell

class Person:
    def __init__(self, first_name, last_name, register):
        self.first_name = first_name
        self.last_name = last_name
        self.register = register
        self.insert_date = datetime.datetime.today()
        self.update_date = datetime.datetime.today()
    def getFirstName(self):
        return self.first_name
    def getLastName(self):
        return self.last_name
    def gerRegister(self):
        return self.register
    def setFirstName(self, first_name):
        self.update_date = datetime.datetime.today()
        self.first_name = first_name
    def setLastName(self, last_name):
        self.update_date = datetime.datetime.today()
        self.last_name = last_name
    def setRegister(self, register):
        self.update_date = datetime.datetime.today()
        self.register = register

class Exam():
    def __init__(self, patientRG, auditorName, auditorStr, weight, height, systolic, diastolic, hemodialysis, dialyses, hemoglobin, albumin, phosphor):
        self.patientRG = patientRG
        self.auditorName = auditorName
        self.auditorStr = auditorStr
        self.weight = float(weight)
        self.height = round(float(int(height)) / 100, 2)
        self.systolic = float(systolic)
        self.diastolic = float(diastolic)
        self.hemodialysis = float(hemodialysis)
        self.dialyses = float(dialyses)
        self.hemoglobin = float(hemoglobin)
        self.albumin = float(albumin)
        self.phosphor = float(phosphor)
        self.insert_date = datetime.datetime.today()
        self.update_date = datetime.datetime.today()     
    def getIMC(self):
        return round(float(self.weight / (self.height * self.height)), 2)
    def getIMCid(self):
        _imc_ = self.getIMC()
        return 0 if _imc_ < 10 else (
            1 if 10 <= _imc_ <= 12.9 else (
                2 if 13 <= _imc_ <= 15.9 else (
                    3 if 16 <= _imc_ <= 16.9 else (
                        4 if 17 <= _imc_ <= 18.4 else (
                            5 if 18.5 <= _imc_ <= 24.9 else (
                                6 if 30 <= _imc_ <= 29.9 else (
                                    7 if 30 <= _imc_ <= 34.5 else (
                                        8 if 35 <= _imc_ <= 39.9 else 9
                                    )
                                )
                            )
                        )
                    )
                )
            )
        )
    def getIMCstr(self):
        _str_ = [ 'Desnutrição grau V', 'Desnutrição grau IV', 'Desnutrição grau III', 'Desnutrição grau II', 'Desnutrição grau I', 'Normal', 'Pré-obesidade', 'Obesidade grau I', 'Obesidade grau II', 'Obesidade grau III' ]
        _imc_ = self.getIMC()
        return _str_[0 if _imc_ < 10 else (
            1 if 10 <= _imc_ <= 12.9 else (
                2 if 13 <= _imc_ <= 15.9 else (
                    3 if 16 <= _imc_ <= 16.9 else (
                        4 if 17 <= _imc_ <= 18.4 else (
                            5 if 18.5 <= _imc_ <= 24.9 else (
                                6 if 30 <= _imc_ <= 29.9 else (
                                    7 if 30 <= _imc_ <= 34.5 else (
                                        8 if 35 <= _imc_ <= 39.9 else 9
                                    )
                                )
                            )
                        )
                    )
                )
            )
        )]
    def getHypID(self):
        s, d = self.systolic, self.diastolic
        return 5 if s > 180 or d > 110 else (
            4 if 180 >= s > 160 or 110 >= d > 100 else (
                3 if 160 >= s > 140 or 100 >= d > 90 else (
                    2 if 140 >= s > 130 or 90 >= d > 85 else (
                        1 if 130 >= s > 120 or 85 >= d > 80 else 0
                    )
                )
            )
        )
    def getHyp(self):
        s, d = self.systolic, self.diastolic
        _str_ = ['Normal (ótimo)', 'Normal', 'Normal (em risco)', 'Hipertensão leve', 'Hipertensão moderada', 'Hipertensão grave']
        return _str_[5 if s > 180 or d > 110 else (
            4 if 180 >= s > 160 or 110 >= d > 100 else (
                3 if 160 >= s > 140 or 100 >= d > 90 else (
                    2 if 140 >= s > 130 or 90 >= d > 85 else (
                        1 if 130 >= s > 120 or 85 >= d > 80 else 0
                    )
                )
            )
        )]

class Patient(Person):
    def __init__(self, first_name, last_name, register, date_birth, genre, date_insurance):
        super(Patient, self).__init__(first_name, last_name, register)
        self.date_birth = date_birth
        self.genre = genre
        self.date_insurance = date_insurance
        self.exams = 0
        self.insert_date = datetime.datetime.today()
        self.update_date = datetime.datetime.today()
    def addExam(self):
        self.exams += 1
        self.update_date = datetime.datetime.today()
    def getAge(self):
        today = datetime.datetime.today()
        y, m, d = [int(x) for x in self.date_birth.split('-')]
        return today.year - y - ((today.month, today.day) < (m, d))

class Patients():
    def __init__(self, data):
        self.n, self.e = -1, -1
        self.patients = []
        self.exams = []
        for r in data[1:]:
            self.addPatient(Patient(
                r[4] + ' ' + r[5], r[6] + ' ' + r[7], r[8], r[9], r[10], r[11], 
            ))
            self.addExam(Exam(
                r[8], r[2], r[3], r[12], r[13], r[14], r[15], r[16], r[17], r[18], r[19], r[20]
            ))
    def addPatient(self, patient):
        self.n += 1
        self.patients.append(patient)
        self.patients.sort(key = lambda x: x.register)
    def addExam(self, exam):
        self.e += 1
        self.exams.append(exam)
    def getExamByID(self, id):
        return self.exams[id]
    def getPatientByID(self, id):
        return self.patients[id]
    def searchPatientID(self, register):
        left = 0
        right = self.n
        while (right >= left):
            middle = int(left + (right - left) / 2) #overflow protection
            u = self.patients[middle]
            if u.register < register:
                left = middle + 1
            elif u.register > register:
                right = middle - 1
            elif u.register == register:
                return middle
        return -1
    def searchPatient(self, register):
        return self.getPatientByID(self.searchPatientID(register))
    def printTable(self, peoplesName, examsName, summaryName):
        file_1 = Workbook()
        file_2 = Workbook()
        file_3 = Workbook()
        peoples = file_1.active
        exams = file_2.active
        summary = file_3.active
        imc_data = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        hyp_data = [0, 0, 0, 0, 0, 0]
        peoples.append((
            'ID', 'Nome', 'Sobrenome', 'RG', 'Sexo', 'Data de nascimento', 'Idade', 'Número de exames', 'Data de entrada no seguro'
        ))
        exams.append((
            'ID', 'Paciente ID', 'Auditor', 'Cargo do auditor', 'Peso (KG)', 'Altura (M)', 'IMC', 'Estado nutricional',
            'Estado de Hipertensão', 'Hemoglobina', 'Albúmina sérica', 'Fósforo'
        ))
        for i, u in enumerate(self.patients):
            peoples.append((
                i, 
                u.first_name, 
                u.last_name, 
                u.register, 
                u.genre, 
                u.date_birth, 
                u.getAge(), 
                u.exams, 
                u.date_insurance
            ))

        for i, e in enumerate(self.exams):
            imc = e.getIMCstr()
            hyp = e.getHyp()
            patient = self.searchPatient(e.patientRG)
            imc_data[e.getIMCid()] += 1
            hyp_data[e.getHypID()] += 1
            exams.append((
                i,
                self.searchPatientID(e.patientRG),
                e.auditorName,
                e.auditorStr,
                e.weight,
                e.height,
                e.getIMC(),
                imc,
                hyp,
                e.hemoglobin,
                e.albumin,
                e.phosphor
            ))
            c = exams.cell(column=8, row=i + 2)
            if imc is not 'Normal':
                c.font = Font(bold=True, name='Arial', color='FFF00000')
                c.fill = PatternFill(fill_type='solid', start_color='ffcccc', end_color='ffcccc')
            else:
                c.font = Font(bold=True, name='Arial', color='006600')
                c.fill = PatternFill(fill_type='solid', start_color='ccffcc', end_color='ccffcc')
            c = exams.cell(column=9, row=i + 2)
            if 'Normal (ótimo)' in hyp:
                c.font = Font(bold=True, name='Arial', color='006600')
                c.fill = PatternFill(fill_type='solid', start_color='ccffcc', end_color='ccffcc')
            elif 'Normal (em risco)' in hyp:
                c.font = Font(name='Arial', color='ff7500')
                c.fill = PatternFill(fill_type='solid', start_color='fff3df', end_color='fff3df')
            elif 'Normal' in hyp:
                c.font = Font(name='Arial', color='006600')
                c.fill = PatternFill(fill_type='solid', start_color='ccffcc', end_color='ccffcc')
            elif 'Hipertensão leve' in hyp:
                c.font = Font(bold=True, name='Arial', color='FFF00000')
                c.fill = PatternFill(fill_type='solid', start_color='ffcccc', end_color='ffcccc')
            elif 'Hipertensão moderada' in hyp:
                c.font = Font(bold=True, name='Arial', color='FFF00000')
                c.fill = PatternFill(fill_type='solid', start_color='ffcccc', end_color='ffcccc')
            elif 'Hipertensão grave' in hyp:
                c.font = Font(bold=True, name='Arial', color='FFFFFFFF')
                c.fill = PatternFill(fill_type='solid', start_color='cc0000', end_color='cc0000')
            c = exams.cell(column=10, row=i + 2)
            if patient.genre is 'M':
                if e.hemoglobin < 14:
                    c.font = Font(bold=True, name='Arial', color='FFF00000')
                    c.fill = PatternFill(fill_type='solid', start_color='ffcccc', end_color='ffcccc')
                elif 14 <= e.hemoglobin <= 18:
                    c.font = Font(bold=True, name='Arial', color='006600')
                    c.fill = PatternFill(fill_type='solid', start_color='ccffcc', end_color='ccffcc')
                else:
                    c.font = Font(bold=True, name='Arial', color='FFFFFFFF')
                    c.fill = PatternFill(fill_type='solid', start_color='cc0000', end_color='cc0000')
            else:
                if e.hemoglobin < 12:
                    c.font = Font(bold=True, name='Arial', color='FFF00000')
                    c.fill = PatternFill(fill_type='solid', start_color='ffcccc', end_color='ffcccc')
                elif 12 <= e.hemoglobin <= 16:
                    c.font = Font(bold=True, name='Arial', color='006600')
                    c.fill = PatternFill(fill_type='solid', start_color='ccffcc', end_color='ccffcc')
                else:
                    c.font = Font(bold=True, name='Arial', color='FFFFFFFF')
                    c.fill = PatternFill(fill_type='solid', start_color='cc0000', end_color='cc0000')
            c = exams.cell(column=11, row=i + 2)
            if 3.5 <= e.albumin <= 5:
                c.font = Font(bold=True, name='Arial', color='FFF00000')
                c.fill = PatternFill(fill_type='solid', start_color='ffcccc', end_color='ffcccc')
            else:
                c.font = Font(bold=True, name='Arial', color='006600')
                c.fill = PatternFill(fill_type='solid', start_color='ccffcc', end_color='ccffcc')
            c = exams.cell(column=12, row=i + 2)
            if 2.5 <= e.phosphor <= 4.5:
                c.font = Font(bold=True, name='Arial', color='FFF00000')
                c.fill = PatternFill(fill_type='solid', start_color='ffcccc', end_color='ffcccc')
            else:
                c.font = Font(bold=True, name='Arial', color='006600')
                c.fill = PatternFill(fill_type='solid', start_color='ccffcc', end_color='ccffcc')

        exams.append((
            '',
            'MÉDIA',
            '',
            '',
            '=MEDIAN(E1:E43)',
            '=MEDIAN(F1:F43)',
            '=MEDIAN(G1:G43)',
            '',
            '',
            '=MEDIAN(J1:J43)',
            '=MEDIAN(K1:K43)',
            '=MEDIAN(L1:L43)'
        ))
        exams.append((
            '',
            'DESVIO PADRÃO',
            '',
            '',
            '=STDEV(E1:E43)',
            '=STDEV(F1:F43)',
            '=STDEV(G1:G43)',
            '',
            '',
            '=STDEV(J1:J43)',
            '=STDEV(K1:K43)',
            '=STDEV(L1:L43)'	
        ))
        for cell in peoples[1]:
            cell.fill = PatternFill(start_color="aabedd", end_color="aabedd", fill_type = "solid")
            cell.font = Font(bold=True, name='Arial', color='FFFFFFFF')

        for cell in exams[1]:
            cell.fill = PatternFill(start_color="aabedd", end_color="aabedd", fill_type = "solid")
            cell.font = Font(bold=True, name='Arial', color='FFFFFFFF')
                
        for cell in exams[44]:
            cell.fill = PatternFill(start_color="aabedd", end_color="aabedd", fill_type = "solid")
            cell.font = Font(bold=True, name='Arial', color='FFFFFFFF')
        for cell in exams[45]:
            cell.fill = PatternFill(start_color="aabedd", end_color="aabedd", fill_type = "solid")
            cell.font = Font(bold=True, name='Arial', color='FFFFFFFF')

        file_1.save(peoplesName)
        file_2.save(examsName)

        _str_ = [ 'Desnutrição grau V', 'Desnutrição grau IV', 'Desnutrição grau III', 'Desnutrição grau II', 'Desnutrição grau I', 'Normal', 'Pré-obesidade', 'Obesidade grau I', 'Obesidade grau II', 'Obesidade grau III' ]
        summary.append((
            'Estado nutricional', 'Número de pacientes'
        ))
        max, id = 0, 0
        for i, x in enumerate(_str_):
            summary.append((
                x, imc_data[i]
            ))
            if imc_data[i] > max:
                id = i
                max = imc_data[i]
        pie = PieChart()
        labels = Reference(summary, min_col=1, min_row=2, max_row=11)
        data = Reference(summary, min_col=2, min_row=1, max_row=11)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Distribuição nutricional"
        slice = DataPoint(idx=id, explosion=10)
        pie.series[0].data_points = [slice]
        pie.height = 6
        summary.add_chart(pie, "C1")

        summary.append([('')])
        summary.append([('')])
        summary.append((
            'Hipertensão', 'Número de pacientes'
        ))

        _str_ = ['Normal (ótimo)', 'Normal', 'Normal (em risco)', 'Hipertensão leve', 'Hipertensão moderada', 'Hipertensão grave']
        max, id = 0, 0
        for i, x in enumerate(_str_):
            summary.append((
                x, hyp_data[i]
            ))
            if hyp_data[i] > max:
                id = i
                max = hyp_data[i]
        pie2 = PieChart()
        labels = Reference(summary, min_col=1, min_row=15, max_row=20)
        data = Reference(summary, min_col=2, min_row=14, max_row=20)
        pie2.add_data(data, titles_from_data=True)
        pie2.set_categories(labels)
        pie2.title = "Distribuição de hipertensão"
        slice = DataPoint(idx=id, explosion=10)
        pie2.series[0].data_points = [slice]
        pie2.height = 6
        summary.add_chart(pie2, "C14")

        for cell in summary[1]:
            cell.fill = PatternFill(start_color="aabedd", end_color="aabedd", fill_type = "solid")
            cell.font = Font(bold=True, name='Arial', color='FFFFFFFF')
        for cell in summary[14]:
            cell.fill = PatternFill(start_color="aabedd", end_color="aabedd", fill_type = "solid")
            cell.font = Font(bold=True, name='Arial', color='FFFFFFFF')
        file_3.save(summaryName)

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
        
if __name__ == "__main__":
    print(
        bcolors.BOLD + '|=======================================================|\n' + bcolors.ENDC +
        bcolors.BOLD + bcolors.OKGREEN + '\tSIG - Sistema de Informações gerenciais.\n' + bcolors.ENDC +
        bcolors.BOLD + '|=======================================================|\n'    
    )    
    # filename = input()
    wb = load_workbook('DadosSIG.xlsx')
    ws = wb.get_sheet_by_name('Hoja1')
    all_rows = []
    for row in ws:
        current_row = []
        for cell in row:
            current_row.append(cell.value)
        all_rows.append(current_row)

    patients = Patients(all_rows)
    patients.printTable('Pessoas.xlsx', 'Ficha.xlsx', 'Relatório.xlsx')

    ### Criacao das tabelas e preenchimento do banco de dados ### 

    TabelaPessoa = bdclasses.CreateTablePerson()
    TabelaFicha = bdclasses.CreateTableRecord()
    Executar = bdclasses.SQLCommand()

    Executar.execute(TabelaPessoa.getCommand())
    Executar.execute(TabelaFicha.getCommand())

    wb = bdclasses.load_workbook('Pessoas.xlsx')
    ws = wb.get_sheet_by_name('Sheet')
    all_rows = []

    for row in ws:
        current_row = []
        for cell in row:
            current_row.append(cell.value)
        all_rows.append(current_row)

    Preencher_pessoa = bdclasses.FillPersonTable()
    Executar.executemany(Preencher_pessoa.getCommand(), all_rows[1:])
    
    wb = bdclasses.load_workbook('Ficha.xlsx')
    ws = wb.get_sheet_by_name('Sheet')
    all_rows = []

    for row in ws:
        current_row = []
        for cell in row:
            current_row.append(cell.value)
        all_rows.append(current_row)

    Preencher_ficha = bdclasses.FillRecordTable()
    Executar.executemany(Preencher_ficha.getCommand(), all_rows[1 : -2])

    bdclasses.init()
    
    aux = 'Dados inseridos com sucesso'
    print('\n' + bdclasses.Fore.GREEN + bdclasses.Back.RED + '{0:^80}'.format(aux) + bdclasses.Style.RESET_ALL + '\n')

