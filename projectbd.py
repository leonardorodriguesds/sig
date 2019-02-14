import sqlite3
import openpyxl
from openpyxl import Workbook, load_workbook

class SQLCommand:
    def __init__(self):
        self.bd_name = 'data.db'
    def connect(self):
        self._conn = sqlite3.connect(self.bd_name)
    def desconnect(self):
        self._conn.close()
    def execute(self, command):
        self.connect()
        self._cursor = self._conn.cursor()
        self._cursor.execute(command)
        self.desconnect()
    def executemany(self, command, rows_list):
        self.connect()
        self._cursor = self._conn.cursor()
        self._cursor.executemany(command, rows_list)
        self._conn.commit()
        self.desconnect()

class CreateTablePerson:
    def __init__(self):
        self.command = """
        CREATE TABLE IF NOT EXISTS person (
            id INTEGER NOT NULL,
            first_name TEXT NOT NULL,
            second_name TEXT NOT NULL,
            register TEXT NOT NULL PRIMARY KEY,
            genre VARCHAR(1) NOT NULL,
            date_birth DATE NOT NULL,
            age INTEGER NOT NULL,
            number INTEGER DEFAULT 0,
            date_insurance DATE NOT NULL
        );
        """
    def getCommand(self):
        return self.command

class CreateTableRecord:
    def __init__(self):
        self.command = """
        CREATE TABLE IF NOT EXISTS record (
            id INTEGER NOT NULL PRIMARY KEY,
            pacient_id INTEGER NOT NULL,
            auditor TEXT NOT NULL,
            auditor_position TEXT NOT NULL,
            weight FLOAT NOT NULL,
            height FLOAT NOT NULL,
            imc FLOAT NOT NULL,
            nutritional_state TEXT NOT NULL,
            hypertension_state TEXT NOT NULL,
            hemoglobin FLOAT NOT NULL,
            albumin FLOAT NOT NULL,
            phosphor FLOAT NOT NULL,
            FOREIGN KEY(pacient_id) REFERENCES person(id)
        );
        """
    def getCommand(self):
        return self.command

class FillPersonTable:
    def __init__(self):
        self.command = """
        INSERT INTO person (id, first_name, second_name, register, genre, date_birth, age, number, date_insurance) 
        VALUES (?,?,?,?,?,?,?,?,?)
        """
    def getCommand(self):
        return self.command

class FillRecordTable:
    def __init__(self):
        self.command = """
        INSERT INTO record (id, pacient_id, auditor, auditor_position, weight, height, imc, nutritional_state, hypertension_state, hemoglobin, albumin, phosphor) 
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """
    def getCommand(self):
        return self.command


if __name__ == "__main__":
    TabelaPessoa = CreateTablePerson()
    TabelaFicha = CreateTableRecord()
    Executar = SQLCommand()

    Executar.execute(TabelaPessoa.getCommand())
    Executar.execute(TabelaFicha.getCommand())

    wb = load_workbook('Pessoas.xlsx')
    ws = wb.get_sheet_by_name('Sheet')
    all_rows = []

    for row in ws:
        current_row = []
        for cell in row:
            current_row.append(cell.value)
        all_rows.append(current_row)

    Preencher_pessoa = FillPersonTable()
    Executar.executemany(Preencher_pessoa.getCommand(), all_rows[1:])
    
    wb = load_workbook('Ficha.xlsx')
    ws = wb.get_sheet_by_name('Sheet')
    all_rows = []

    for row in ws:
        current_row = []
        for cell in row:
            current_row.append(cell.value)
        all_rows.append(current_row)

    Preencher_ficha = FillRecordTable()
    Executar.executemany(Preencher_ficha.getCommand(), all_rows[1 : -2])

    print('Dados inseridos com sucesso.')
